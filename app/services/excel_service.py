"""
Excel 导入/导出服务
"""
import json
import os
from datetime import date
from app.services import bom_service, part_service
from app.utils.db import get_session
from app.utils.config_manager import ConfigManager
from app.core.audit import write_audit


def export_bom_to_excel(project_id: str, filepath: str):
    """
    导出 BOM 为 Excel，格式与原始模板一致：
    - 顶部：变更履历表（最新3条）
    - 头部：BOM编号、客户零件号、版本、日期
    - 明细：层级标记/标志/零件图号/版本/标准/名称/规格/单位/用量/备注
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    cfg = ConfigManager()
    use_snapshot = cfg.getbool("export", "use_snapshot", True)
    include_changelog = cfg.getbool("export", "include_change_log", True)

    proj = bom_service.get_project(project_id)
    if not proj:
        raise ValueError("项目不存在")

    items = bom_service.get_items(project_id)
    logs = bom_service.get_change_logs(project_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = proj.bom_number[:31]  # sheet名最长31字符

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    title_fill  = PatternFill("solid", fgColor="1F3864")
    title_font  = Font(bold=True, color="FFFFFF", size=11)
    bold        = Font(bold=True)

    row = 1

    # ── 变更履历表 ──
    if include_changelog:
        ws.merge_cells(f"A{row}:N{row}")
        c = ws.cell(row=row, column=1, value="变更履历表")
        c.font = title_font; c.fill = title_fill; c.alignment = Alignment(horizontal="center")
        row += 1

        log_headers = ["序目", "变更内容", "", "", "", "", "变更原因", "", "变更前版本", "变更后版本", "变更人", "变更日期", "备注", ""]
        for col_i, h in enumerate(log_headers, 1):
            cell = ws.cell(row=row, column=col_i, value=h)
            cell.font = bold; cell.fill = header_fill; cell.border = border
        row += 1

        recent_logs = logs[:3]
        for log in recent_logs:
            ws.cell(row=row, column=1, value=log.sequence)
            ws.cell(row=row, column=2, value=log.change_description or "")
            ws.merge_cells(f"B{row}:F{row}")
            ws.cell(row=row, column=7, value=log.change_reason or "")
            ws.merge_cells(f"G{row}:H{row}")
            ws.cell(row=row, column=9, value=log.previous_version or "")
            ws.cell(row=row, column=10, value=log.new_version)
            ws.cell(row=row, column=11, value=log.changed_by_name or "")
            ws.cell(row=row, column=12, value=log.change_date)
            ws.cell(row=row, column=13, value=log.notes or "")
            row += 1

        row += 1  # 空行

    # ── BOM 头部信息 ──
    ws.cell(row=row, column=1, value="BOM编号").font = bold
    ws.cell(row=row, column=2, value=proj.bom_number)
    ws.cell(row=row, column=7, value="客户零件号").font = bold
    ws.cell(row=row, column=8, value=proj.customer_part_number or "")
    ws.merge_cells(f"H{row}:K{row}")
    ws.cell(row=row, column=12, value="建立日期").font = bold
    ws.cell(row=row, column=13, value=proj.established_date or "")
    row += 1

    ws.cell(row=row, column=1, value=proj.bom_number)
    ws.cell(row=row, column=7, value=proj.customer_description or "")
    ws.merge_cells(f"G{row}:K{row}")
    ws.cell(row=row, column=12, value="更新日期").font = bold
    ws.cell(row=row, column=13, value=proj.updated_date or "")
    row += 1

    ws.cell(row=row, column=1, value="零件图号").font = bold
    ws.cell(row=row, column=2, value=proj.bom_number)
    ws.cell(row=row, column=12, value="版本").font = bold
    ws.cell(row=row, column=13, value=proj.current_version)
    row += 2  # 空行

    # ── 明细列标题 ──
    detail_headers = ["标记", "层级", "序号", "", "", "零件图号", "零件版本", "零件标准", "零件名称", "规格/材料", "单位", "用量", "备注"]
    col_widths_map  = {1:6, 2:6, 3:8, 6:18, 7:8, 8:8, 9:22, 10:36, 11:6, 12:6, 13:20}
    for col_i, h in enumerate(detail_headers, 1):
        cell = ws.cell(row=row, column=col_i, value=h)
        cell.font = bold; cell.fill = header_fill; cell.border = border
        cell.alignment = Alignment(horizontal="center")
        if col_i in col_widths_map:
            ws.column_dimensions[get_column_letter(col_i)].width = col_widths_map[col_i]
    row += 1

    # ── 明细数据行 ──
    seq_counters = [0] * 6
    for item in items:
        level = item.level
        seq_counters[level] += 1
        for i in range(level + 1, 6):
            seq_counters[i] = 0
        seq = ".".join(str(seq_counters[i]) for i in range(1, level + 1))

        snap = {}
        if use_snapshot and item.part_snapshot:
            try:
                snap = json.loads(item.part_snapshot)
            except Exception:
                pass
        if not snap:
            part = part_service.get_part_by_number(item.part_number)
            if part:
                snap = part_service.get_part_snapshot(item.part_number)

        level_label = ["", "一", "二", "三", "四", "五"][min(level, 5)]

        row_data = [
            item.level_mark or "",
            level_label,
            seq, "", "",
            item.part_number,
            snap.get("version", ""),
            snap.get("standard_level", ""),
            snap.get("name", ""),
            snap.get("spec_material", ""),
            snap.get("unit", ""),
            item.quantity,
            item.notes or "",
        ]
        for col_i, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_i, value=val)
            cell.border = border
            if col_i in (1, 2, 3, 7, 8, 11, 12):
                cell.alignment = Alignment(horizontal="center")
            # 名称列缩进
            if col_i == 9 and level > 1:
                cell.alignment = Alignment(indent=level - 1)
        row += 1

    os.makedirs(os.path.dirname(filepath) if os.path.dirname(filepath) else ".", exist_ok=True)
    wb.save(filepath)
    write_audit("EXPORT", resource_type="BOM", resource_id=project_id,
                detail=f"导出 BOM {proj.bom_number}: {filepath}")


def import_bom_from_excel(filepath: str) -> dict:
    """
    从 Excel 导入 BOM 明细。
    返回预检结果字典：
    {
      "bom_number": str,
      "version": str,
      "items": [dict, ...],
      "warnings": [str, ...],  # 陌生零件等警告
      "errors": [str, ...],    # 格式错误
    }
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)

    result = {"bom_number": "", "version": "", "items": [], "warnings": [], "errors": []}

    # 取第一个非索引 sheet（通常是项目 sheet）
    sheet_names = [n for n in wb.sheetnames if n not in ("零部件总览表",)]
    if not sheet_names:
        result["errors"].append("未找到有效的 BOM 工作表")
        return result

    ws = wb[sheet_names[0]]
    rows = list(ws.iter_rows(values_only=True))

    # 扫描头部信息（BOM编号 在某行 A 列）
    for i, row in enumerate(rows):
        if row and str(row[0] or "").strip() == "BOM编号":
            result["bom_number"] = str(row[1] or "")
        if row and str(row[11] or "").strip() == "版本":
            result["version"] = str(row[12] or "")
        # 找到明细列标题行
        if row and str(row[5] or "").strip() == "零件图号":
            data_start = i + 1
            break
    else:
        result["errors"].append("未找到明细列标题行（含'零件图号'的行）")
        return result

    for i, row in enumerate(rows[data_start:], data_start):
        part_no = str(row[5] or "").strip()
        if not part_no:
            continue

        level_mark  = str(row[0] or "").strip()
        level_label = str(row[1] or "").strip()
        qty_raw = row[11]
        try:
            qty = float(qty_raw) if qty_raw else 1.0
        except (ValueError, TypeError):
            qty = 1.0

        level_map = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "": 1}
        level = level_map.get(level_label, 1)

        # 校验零件是否在母表中
        part = part_service.get_part_by_number(part_no)
        if not part:
            result["warnings"].append(f"行 {i+1}: 零件 [{part_no}] 不在母表中")

        result["items"].append({
            "sort_order": len(result["items"]),
            "level": level,
            "level_mark": level_mark,
            "level_label": level_label,
            "part_number": part_no,
            "quantity": qty,
            "notes": str(row[12] or "").strip(),
        })

    return result
