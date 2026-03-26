"""
零件母表服务：增删改查 + 导出
"""
from typing import Optional
from app.models import PartsMaster
from app.utils.db import get_session
from app.core.rbac import require_role
from app.core.audit import write_audit
from app.core.auth import AuthContext


def search_parts(keyword: str = "", include_deleted: bool = False) -> list[PartsMaster]:
    with get_session() as s:
        q = s.query(PartsMaster)
        if not include_deleted:
            q = q.filter_by(is_deleted=0)
        if keyword:
            kw = f"%{keyword}%"
            from sqlalchemy import or_
            q = q.filter(or_(
                PartsMaster.part_number.ilike(kw),
                PartsMaster.name.ilike(kw),
                PartsMaster.spec_material.ilike(kw),
            ))
        return q.order_by(PartsMaster.part_number).all()


def get_part_by_number(part_number: str) -> Optional[PartsMaster]:
    with get_session() as s:
        return s.query(PartsMaster).filter_by(
            part_number=part_number, is_deleted=0
        ).first()


def get_part_snapshot(part_number: str) -> dict:
    """返回零件当前数据的快照字典，用于写入 BomItem.part_snapshot"""
    part = get_part_by_number(part_number)
    if not part:
        return {"part_number": part_number}
    return {
        "part_number": part.part_number,
        "version": part.version,
        "standard_level": part.standard_level,
        "name": part.name,
        "spec_material": part.spec_material,
        "unit": part.unit,
        "default_qty": part.default_qty,
        "notes": part.notes,
    }


@require_role("super_admin")
def create_part(
    part_number: str, version: str, name: str,
    standard_level: str = None, spec_material: str = None,
    unit: str = "pcs", default_qty: float = 1.0, notes: str = None,
) -> PartsMaster:
    existing = get_part_by_number(part_number)
    if existing:
        raise ValueError(f"零件图号 '{part_number}' 已存在")

    user_id = AuthContext().user.id if AuthContext().user else None
    part = PartsMaster(
        part_number=part_number, version=version, name=name,
        standard_level=standard_level, spec_material=spec_material,
        unit=unit, default_qty=default_qty, notes=notes,
        created_by=user_id, updated_by=user_id,
    )
    with get_session() as s:
        s.add(part)
        s.commit()
        s.refresh(part)

    write_audit("CREATE", resource_type="PART", resource_id=part_number,
                data_after={"part_number": part_number, "name": name})
    return part


@require_role("super_admin")
def update_part(part_id: int, **fields) -> PartsMaster:
    allowed = {"version", "name", "standard_level", "spec_material", "unit", "default_qty", "notes"}
    with get_session() as s:
        part = s.get(PartsMaster, part_id)
        if not part or part.is_deleted:
            raise ValueError("零件不存在")

        before = {k: getattr(part, k) for k in allowed if hasattr(part, k)}
        for k, v in fields.items():
            if k in allowed:
                setattr(part, k, v)
        part.updated_by = AuthContext().user.id if AuthContext().user else None
        s.commit()
        s.refresh(part)

    write_audit("UPDATE", resource_type="PART", resource_id=str(part_id),
                data_before=before,
                data_after={k: getattr(part, k) for k in allowed})
    return part


@require_role("super_admin")
def soft_delete_part(part_id: int):
    with get_session() as s:
        part = s.get(PartsMaster, part_id)
        if not part or part.is_deleted:
            raise ValueError("零件不存在")
        before = {"part_number": part.part_number, "name": part.name}
        part.soft_delete()
        s.commit()

    write_audit("SOFT_DELETE", resource_type="PART", resource_id=str(part_id),
                data_before=before)


@require_role("super_admin")
def import_parts_from_excel(filepath: str) -> tuple[int, int, list[str]]:
    """
    从 Excel 批量导入零件母表。
    列顺序（第1行为表头，忽略；从第2行开始读）：
      A: 零件图号  B: 版本  C: 零件标准  D: 零件名称
      E: 规格/材料  F: 单位  G: 用量  H: 备注

    Returns: (added_count, skipped_count, error_msgs)
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    user_id = AuthContext().user.id if AuthContext().user else None
    added = 0
    skipped = 0
    errors: list[str] = []

    with get_session() as s:
        for i, row in enumerate(rows, start=2):
            if not row or not row[0]:
                continue
            part_number = str(row[0]).strip()
            if not part_number:
                continue

            version      = str(row[1]).strip() if row[1] is not None else ""
            standard     = str(row[2]).strip() if row[2] is not None else None
            name         = str(row[3]).strip() if row[3] is not None else ""
            spec         = str(row[4]).strip() if row[4] is not None else None
            unit         = str(row[5]).strip() if row[5] is not None else "pcs"
            try:
                qty = float(row[6]) if row[6] is not None else 1.0
            except (ValueError, TypeError):
                qty = 1.0
            notes        = str(row[7]).strip() if row[7] is not None else None

            if not version:
                errors.append(f"第{i}行 [{part_number}]：版本号为空，已跳过")
                skipped += 1
                continue
            if not name:
                errors.append(f"第{i}行 [{part_number}]：零件名称为空，已跳过")
                skipped += 1
                continue

            existing = s.query(PartsMaster).filter_by(part_number=part_number, is_deleted=0).first()
            if existing:
                skipped += 1
                continue

            part = PartsMaster(
                part_number=part_number, version=version,
                standard_level=standard or None, name=name,
                spec_material=spec, unit=unit or "pcs",
                default_qty=qty, notes=notes,
                created_by=user_id, updated_by=user_id,
            )
            s.add(part)
            added += 1

        s.commit()

    write_audit("IMPORT", resource_type="PART",
                detail=f"从 Excel 导入零件：新增 {added} 条，跳过 {skipped} 条，错误 {len(errors)} 条")
    return added, skipped, errors


def export_parts_to_excel(filepath: str):
    """导出零件母表为 Excel，格式与原始文件一致（8列）"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    parts = search_parts()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "零部件总览表"

    headers = ["零件图号", "零件版本", "零件标准", "零件名称", "规格/材料", "单位", "用量", "备注"]
    col_widths = [18, 10, 10, 20, 36, 8, 8, 20]

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    for row, p in enumerate(parts, 2):
        row_data = [p.part_number, p.version, p.standard_level, p.name,
                    p.spec_material, p.unit, p.default_qty, p.notes]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border

    wb.save(filepath)
    write_audit("EXPORT", resource_type="PART", detail=f"导出零件母表: {filepath}")
